import uuid
import json
import hashlib
from datetime import datetime, timezone, date as date_type
from io import BytesIO

import requests

from peptide_radar.utils.diff_engine import diff_structured_rows
from peptide_radar.resolvers.entity_resolver import resolve_peptide, build_alias_index, normalize

SOURCES = {
    "fda_503a": {
        "url": "https://www.fda.gov/media/94155/download",
        "content_type": "xlsx",
    },
    "fda_503b": {
        "url": "https://www.fda.gov/drugs/human-drug-compounding/503b-bulk-drug-substances-list",
        "content_type": "html",
    },
    "fda_safety_risk": {
        "url": "https://www.fda.gov/drugs/human-drug-compounding/certain-bulk-drug-substances-use-compounding-may-present-significant-safety-risks",
        "content_type": "html",
    },
}


def fetch_source(url):
    """Fetch raw bytes from URL. Never raises."""
    try:
        resp = requests.get(url, timeout=60)
        return resp.content, resp.status_code
    except Exception as e:
        print(f"  fetch_source error: {e}")
        return b"", 0


def _snapshot_hash(raw_bytes):
    return hashlib.sha256(raw_bytes).hexdigest()


def _get_prior_snapshot(source):
    """Returns (prior_hash, prior_raw_content) or (None, None)."""
    try:
        row = spark.sql(f"""
            SELECT snapshot_id, raw_content
            FROM peptide_radar.bronze.raw_snapshots
            WHERE source = '{source}'
            ORDER BY fetch_timestamp DESC
            LIMIT 1
        """).first()
        if row:
            raw = bytes(row["raw_content"]) if row["raw_content"] else None
            return row["snapshot_id"], raw
        return None, None
    except NameError:
        return None, None


def _write_snapshot(snapshot_id, source, source_url, raw_content, content_type,
                    prior_hash, changed, http_status):
    try:
        data = [{
            "snapshot_id": snapshot_id,
            "source": source,
            "source_url": source_url,
            "raw_content": bytearray(raw_content),
            "content_type": content_type,
            "fetch_timestamp": datetime.now(timezone.utc),
            "prior_hash": prior_hash,
            "changed": changed,
            "http_status": http_status,
            "parser_version": "1.0",
        }]
        spark.createDataFrame(data).write.mode("append").saveAsTable(
            "peptide_radar.bronze.raw_snapshots"
        )
    except NameError:
        pass


def parse_503a(raw_bytes):
    """Parse 503A Excel file into structured rows."""
    rows = []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(raw_bytes), read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [str(c.value).strip().lower() if c.value else "" for c in header_row]

        name_col = cat_col = status_col = date_col = None
        for i, h in enumerate(headers):
            if name_col is None and ("substance" in h or "name" in h or "ingredient" in h):
                name_col = i
            elif cat_col is None and "category" in h:
                cat_col = i
            elif status_col is None and "status" in h:
                status_col = i
            elif date_col is None and ("date" in h or "effective" in h):
                date_col = i

        if name_col is None:
            name_col = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            cells = list(row)
            if name_col >= len(cells) or not cells[name_col]:
                continue
            name = str(cells[name_col]).strip()
            if not name or name == "None":
                continue
            rows.append({
                "peptide_name": name,
                "category": str(cells[cat_col]).strip() if cat_col is not None and cat_col < len(cells) and cells[cat_col] else "503a_bulk",
                "status_text": str(cells[status_col]).strip() if status_col is not None and status_col < len(cells) and cells[status_col] else "",
                "effective_date": str(cells[date_col]).strip() if date_col is not None and date_col < len(cells) and cells[date_col] else "",
            })
        wb.close()
    except Exception as e:
        print(f"  parse_503a error: {e}")
    return rows


def parse_503b_safety(raw_bytes, source_type):
    """Parse 503B or safety risk HTML page into structured rows."""
    rows = []
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(raw_bytes, "html.parser")

        for table in soup.find_all("table"):
            trs = table.find_all("tr")
            if len(trs) < 2:
                continue

            headers = [th.get_text(strip=True).lower() for th in trs[0].find_all(["th", "td"])]
            name_col = cat_col = status_col = date_col = None
            for i, h in enumerate(headers):
                if name_col is None and ("substance" in h or "name" in h or "ingredient" in h or "drug" in h):
                    name_col = i
                elif cat_col is None and "category" in h:
                    cat_col = i
                elif status_col is None and ("status" in h or "reason" in h):
                    status_col = i
                elif date_col is None and ("date" in h or "effective" in h):
                    date_col = i

            if name_col is None:
                name_col = 0

            default_cat = "fda_safety_risk" if source_type == "fda_safety_risk" else "503b"

            for tr in trs[1:]:
                cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
                if not cells or name_col >= len(cells):
                    continue
                name = cells[name_col].strip()
                if not name:
                    continue
                rows.append({
                    "peptide_name": name,
                    "category": cells[cat_col].strip() if cat_col is not None and cat_col < len(cells) else default_cat,
                    "status_text": cells[status_col].strip() if status_col is not None and status_col < len(cells) else "",
                    "effective_date": cells[date_col].strip() if date_col is not None and date_col < len(cells) else "",
                })
    except Exception as e:
        print(f"  parse_503b_safety error ({source_type}): {e}")
    return rows


def _load_alias_index():
    try:
        rows = spark.sql("""
            SELECT alias, canonical_name
            FROM peptide_radar.silver.peptide_aliases
        """).collect()
        return build_alias_index([
            {"alias": r["alias"], "canonical_name": r["canonical_name"]} for r in rows
        ])
    except NameError:
        return {}


def _load_peptide_lookup():
    """canonical_name -> peptide_id from silver.peptides."""
    try:
        rows = spark.sql("""
            SELECT peptide_id, canonical_name
            FROM peptide_radar.silver.peptides
        """).collect()
        return {r["canonical_name"]: r["peptide_id"] for r in rows}
    except NameError:
        return {}


def _load_fda_mapping():
    """raw_value -> normalized_status from silver.fda_category_mapping."""
    try:
        rows = spark.sql("""
            SELECT raw_value, normalized_status
            FROM peptide_radar.silver.fda_category_mapping
        """).collect()
        return {r["raw_value"]: r["normalized_status"] for r in rows}
    except NameError:
        return {}


def _resolve_name(peptide_name, alias_index, peptide_lookup):
    """Resolve peptide name to (peptide_id, canonical_name) or (None, None)."""
    resolved = resolve_peptide(peptide_name, alias_index)
    if resolved:
        for canonical in resolved:
            pid = peptide_lookup.get(canonical)
            if pid:
                return pid, canonical
    norm = normalize(peptide_name)
    for canonical, pid in peptide_lookup.items():
        if normalize(canonical) == norm:
            return pid, canonical
    return None, None


def _write_rows(table_name, rows):
    if not rows:
        return
    try:
        spark.createDataFrame(rows).write.mode("append").saveAsTable(table_name)
    except NameError:
        pass


def _cross_check(peptide_id, peptide_name, external_category, fda_mapping):
    """Compare external FDA category against internal data. Returns discrepancy dict or None."""
    try:
        safe_name = peptide_name.replace("'", "''").lower()
        row = spark.sql(f"""
            SELECT fda_approved
            FROM elevanbio_dev.bronze.peptide_database_raw
            WHERE LOWER(peptide_name_generic) = '{safe_name}'
        """).first()
        if not row or not row["fda_approved"]:
            return None

        internal = str(row["fda_approved"]).strip().lower()
        normalized_ext = fda_mapping.get(external_category, "unknown")

        internal_approved = internal.startswith("yes")
        external_approved = normalized_ext == "approved"

        if internal_approved != external_approved:
            return {
                "discrepancy_id": str(uuid.uuid4()),
                "peptide_id": peptide_id,
                "peptide_name": peptide_name,
                "field_name": "fda_category",
                "internal_value": row["fda_approved"],
                "external_value": external_category,
                "external_source": "fda_bulks",
                "severity": "high",
                "detected_at": datetime.now(timezone.utc),
                "reviewed": False,
            }
    except NameError:
        pass
    return None


def process_source(source_name, url, content_type):
    """Process one FDA source end-to-end. Returns result dict."""
    result = {
        "source": source_name,
        "changed": False,
        "baseline": False,
        "signals_written": 0,
        "unresolved_count": 0,
        "discrepancy_count": 0,
        "alerts": [],
    }

    # Step 1: Fetch
    raw_bytes, http_status = fetch_source(url)
    if not raw_bytes:
        print(f"  {source_name}: empty response (status={http_status})")
        return result

    # Step 2: Hash
    current_hash = _snapshot_hash(raw_bytes)

    # Step 3: Compare to prior
    prior_hash, prior_content = _get_prior_snapshot(source_name)
    if current_hash == prior_hash:
        print(f"  {source_name}: no change")
        return result

    result["changed"] = True

    # Step 4: Write snapshot
    _write_snapshot(current_hash, source_name, url, raw_bytes, content_type,
                    prior_hash, True, http_status)

    # First run — baseline only, no signals
    if prior_hash is None:
        print(f"  {source_name}: baseline snapshot stored")
        result["baseline"] = True
        return result

    # Step 5: Parse current and prior
    if content_type == "xlsx":
        new_rows = parse_503a(raw_bytes)
        old_rows = parse_503a(prior_content) if prior_content else []
    else:
        new_rows = parse_503b_safety(raw_bytes, source_name)
        old_rows = parse_503b_safety(prior_content, source_name) if prior_content else []

    # Step 6: Diff
    diff = diff_structured_rows(old_rows, new_rows, "peptide_name", ["category", "status_text"])
    if not diff["changed"] and not diff["inserted"] and not diff["deleted"]:
        print(f"  {source_name}: content changed but no structural differences")
        return result

    # Step 7: Load resolution data
    alias_index = _load_alias_index()
    peptide_lookup = _load_peptide_lookup()
    fda_mapping = _load_fda_mapping()

    signals = []
    unresolved = []
    discrepancies = []
    now = datetime.now(timezone.utc)
    today = date_type.today()

    def make_signal(name, pid, event_type, event_value, direction, severity):
        return {
            "signal_id": str(uuid.uuid4()),
            "peptide_id": pid,
            "event_date": today,
            "source_type": source_name,
            "event_type": event_type,
            "event_value": json.dumps(event_value),
            "event_direction": direction,
            "severity": severity,
            "raw_ref": name,
            "source_hash": current_hash,
            "sent_to_gold": False,
            "signal_date": now,
        }

    def make_review(name):
        return {
            "review_id": str(uuid.uuid4()),
            "raw_text": name,
            "source_type": source_name,
            "source_ref": url,
            "candidate_aliases": [],
            "status": "pending",
            "resolved_to": None,
            "reviewer_note": None,
            "created_at": now,
            "reviewed_at": None,
        }

    # Process changed rows
    for old_row, new_row in diff["changed"]:
        name = new_row["peptide_name"]
        pid, canonical = _resolve_name(name, alias_index, peptide_lookup)
        if pid is None:
            unresolved.append(make_review(name))
            continue

        ev = {
            "old_category": old_row.get("category"),
            "new_category": new_row.get("category"),
            "old_status": old_row.get("status_text"),
            "new_status": new_row.get("status_text"),
        }
        sig = make_signal(name, pid, "fda_category_change", ev, "negative", "high")
        signals.append(sig)
        result["alerts"].append(sig)

        disc = _cross_check(pid, name, new_row.get("category", ""), fda_mapping)
        if disc:
            discrepancies.append(disc)

    # Process inserted rows
    for row in diff["inserted"]:
        name = row["peptide_name"]
        pid, canonical = _resolve_name(name, alias_index, peptide_lookup)
        if pid is None:
            unresolved.append(make_review(name))
            continue

        if source_name == "fda_safety_risk":
            event_type = "fda_safety_risk_added"
            direction = "negative"
            severity = "critical"
        else:
            event_type = "fda_category_change"
            direction = "neutral"
            severity = "medium"

        ev = {"category": row.get("category"), "status_text": row.get("status_text")}
        sig = make_signal(name, pid, event_type, ev, direction, severity)
        signals.append(sig)
        if event_type in ("fda_category_change", "fda_safety_risk_added"):
            result["alerts"].append(sig)

        disc = _cross_check(pid, name, row.get("category", ""), fda_mapping)
        if disc:
            discrepancies.append(disc)

    # Process deleted rows — only signal safety risk removals
    for row in diff["deleted"]:
        if source_name != "fda_safety_risk":
            continue
        name = row["peptide_name"]
        pid, canonical = _resolve_name(name, alias_index, peptide_lookup)
        if pid is None:
            continue
        ev = {"category": row.get("category"), "status_text": row.get("status_text")}
        sig = make_signal(name, pid, "fda_safety_risk_removed", ev, "positive", "medium")
        signals.append(sig)

    # Step 8: Write results
    _write_rows("peptide_radar.silver.signals", signals)
    _write_rows("peptide_radar.silver.manual_review_queue", unresolved)
    _write_rows("peptide_radar.silver.internal_discrepancies", discrepancies)

    result["signals_written"] = len(signals)
    result["unresolved_count"] = len(unresolved)
    result["discrepancy_count"] = len(discrepancies)

    print(f"  {source_name}: {len(signals)} signals, {len(unresolved)} unresolved, {len(discrepancies)} discrepancies")
    return result
